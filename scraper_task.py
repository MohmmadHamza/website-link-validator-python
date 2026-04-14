import os
import re
import time
import argparse
from urllib.parse import urljoin
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}


def read_urls_from_excel(file_path: str):
    wb = load_workbook(file_path)
    ws = wb.active

    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            url = str(row[0]).strip()
            if url:
                urls.append(url)
    return urls


def fetch_html(url: str, timeout: int = 20):
    try:
        response = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        response.raise_for_status()
        return response.text, None
    except requests.RequestException as exc:
        return None, str(exc)


def extract_anchor_links(page_url: str, html: str):
    soup = BeautifulSoup(html, "html.parser")
    links = set()

    for a_tag in soup.find_all("a", href=True):
        href = a_tag.get("href", "").strip()

        if not href:
            continue

        if href.startswith(("javascript:", "mailto:", "tel:", "#")):
            continue

        absolute_link = urljoin(page_url, href)
        links.add(absolute_link)

    return sorted(links)


def extract_image_links(page_url: str, html: str):
    soup = BeautifulSoup(html, "html.parser")
    images = set()

    for img_tag in soup.find_all("img", src=True):
        src = img_tag.get("src", "").strip()
        if not src:
            continue
        absolute_src = urljoin(page_url, src)
        images.add(absolute_src)

    return sorted(images)


def validate_link(url: str, timeout: int = 15):
    try:
        response = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0"
            },
            timeout=timeout,
            allow_redirects=True
        )

        if response.status_code >= 400:
            return True, f"HTTP {response.status_code}"

        return False, f"HTTP {response.status_code}"

    except requests.exceptions.RequestException as e:
        return True, str(e)


def save_broken_links_to_excel(results, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Broken Links"
    ws.append(["pageURL", "Broken Links", "Status"])

    for row in results:
        ws.append([row["pageURL"], row["broken_link"], row["status"]])

    wb.save(output_path)


def sanitize_filename(name: str):
    name = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
    return name[:180] if name else "image"


def guess_extension_from_url(url: str):
    path = url.split("?")[0]
    _, ext = os.path.splitext(path)
    if ext and len(ext) <= 5:
        return ext
    return ".jpg"


def download_images(image_urls, output_dir: str):
    os.makedirs(output_dir, exist_ok=True)
    downloaded = []

    for index, img_url in enumerate(image_urls, start=1):
        try:
            response = requests.get(img_url, headers=HEADERS, timeout=25, stream=True)
            response.raise_for_status()

            filename = os.path.basename(img_url.split("?")[0]).strip()
            if not filename:
                filename = f"image_{index}{guess_extension_from_url(img_url)}"
            else:
                filename = sanitize_filename(filename)

            file_path = os.path.join(output_dir, filename)

            base, ext = os.path.splitext(file_path)
            counter = 1
            while os.path.exists(file_path):
                file_path = f"{base}_{counter}{ext}"
                counter += 1

            with open(file_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)

            downloaded.append({"image_url": img_url, "saved_as": file_path, "status": "Downloaded"})

        except requests.RequestException as exc:
            downloaded.append({"image_url": img_url, "saved_as": "", "status": f"Failed: {exc}"})

    return downloaded


def task_1_find_broken_links(input_excel: str, output_excel: str):
    urls = read_urls_from_excel(input_excel)
    all_broken_links = []

    for page_url in urls:
        print(f"Processing page: {page_url}")
        html, error = fetch_html(page_url)

        if error:
            all_broken_links.append({
                "pageURL": page_url,
                "broken_link": "[Could not load page itself]",
                "status": error
            })
            continue

        extracted_links = extract_anchor_links(page_url, html)
        print(f"  Found {len(extracted_links)} anchor links")

        for link in extracted_links:
            is_broken, status = validate_link(link)
            if is_broken:
                all_broken_links.append({
                    "pageURL": page_url,
                    "broken_link": link,
                    "status": status
                })
            time.sleep(0.2)

    save_broken_links_to_excel(all_broken_links, output_excel)
    print(f"\nBroken links report saved to: {output_excel}")


def task_2_download_images(input_excel: str, output_dir: str):
    urls = read_urls_from_excel(input_excel)
    all_images = set()

    for page_url in urls:
        print(f"Processing page: {page_url}")
        html, error = fetch_html(page_url)

        if error:
            print(f"  Could not load page: {error}")
            continue

        image_links = extract_image_links(page_url, html)
        print(f"  Found {len(image_links)} images")
        all_images.update(image_links)

    print(f"\nTotal unique images to download: {len(all_images)}")
    download_results = download_images(sorted(all_images), output_dir)

    success_count = sum(1 for item in download_results if item["status"] == "Downloaded")
    print(f"Downloaded {success_count} images into: {output_dir}")


def main():
    parser = argparse.ArgumentParser(description="Website link checker and image downloader from Excel URLs")
    parser.add_argument("--task", choices=["broken_links", "download_images"], required=True, help="Task to run")
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output", required=True, help="Output Excel path (task 1) or output folder (task 2)")
    args = parser.parse_args()

    if args.task == "broken_links":
        task_1_find_broken_links(args.input, args.output)
    elif args.task == "download_images":
        task_2_download_images(args.input, args.output)


if __name__ == "__main__":
    main()
